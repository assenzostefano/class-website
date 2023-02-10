from bson.objectid import ObjectId
from openpyxl.styles import NamedStyle
from dotenv import load_dotenv
import openpyxl as xl
import datetime
import pymongo
import urllib
import datetime
import os

load_dotenv() #Load .env file
PASSWORD_MONGODB = os.getenv('PASSWORD_MONGODB') #Password for MongoDB
URL_MONGODB = os.getenv('URL_MONGODB') #URL for MongoDB
mongo_url = "mongodb+srv://elci:" + urllib.parse.quote_plus(PASSWORD_MONGODB) + URL_MONGODB #URL for MongoDB (with password)
client = pymongo.MongoClient(mongo_url) #Connect to MongoDB
database = client["website-class"] #Database name
collection = database["school-time-table"] #Collection school time table current
collection_archive = database["archive-school-time-table"] #Collection school time table archive

def update_time_school():
    # Load excel file
    namefile_xlsx = "attachments/school_time.xlsx"
    workbook = xl.load_workbook(filename=namefile_xlsx)
    date_style = NamedStyle(name='date_style',number_format='dd/mm/yy')
    workbook.add_named_style(date_style)
    ws = workbook.active

    #Date
    current_time = datetime.datetime.now()
    day = str(current_time.day)
    month = str(current_time.month)
    year = str(current_time.year)
    hour = str(current_time.hour)
    minute = str(current_time.minute)
    long_date = day + "-" + month + "-" + year + " " + hour + ":" + minute

    #Create collection
    mydict = {
        "Date": long_date, 
        "School Subject": {
            "Monday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
            "Tuesday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
            "Wednesday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
            "Thursday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
            "Friday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
            "Saturday": [
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
                {
                    "Subject": "",
                    "Teacher": "",
                    "Room": "",
                    "PE with": "",
                    "le is busy": "",
                    "le3 is busy": "",
                },
            ],
    }
}
    x = collection.delete_many({}) #Delete all documents in collection (school-time-table)
    x = collection.insert_one(mydict) # Add collection on collection (school-time-table)
    x = collection_archive.insert_one(mydict) # Add collection on collection (archive-school-time-table)
    check_repeat = 0
    check_repeat_teacher = 0
    check_repeat_room = 0
    dont_repeat = 0
    dont_repeat_teacher = 0
    dont_repeat_room = 0
    gagaga_teacher = 0
    gagaga_room = 0
    number_teacher = 1
    number_room = 1
    number = 1
    current_day = None
    day_counter = 0
    number_day = 0
    #Search my class in excel file and add in MongoDB
    number = 1
    gagaga = 0
    for row in range (1, 100):
        # column B ~ column F
        for column in range (1, 100):
            cell = ws.cell(row, column)
            if cell.value == "2elci":
                #Search school time table
                for i in range(4,80):
                    day = str(ws.cell(row=i, column=3).value) # Get day from excel file
                    school_subject = ws.cell(row=i, column=column).value # Get school subject from excel file
                    teacher = ws.cell(row=i, column=column+1).value
                    room = ws.cell(row=i, column=column+2).value
                    if dont_repeat == 9:
                        check_repeat += 1
                        if check_repeat == 5:
                            check_repeat = 0
                            dont_repeat = 0
                    else:
                        if day == "None":
                            if school_subject == 0: #If school subject is 0, add "" in MongoDB
                                number += 1
                                gagaga += 1
                                dont_repeat += 1
                                if number == 9:
                                    number = 1
                                if gagaga == 9:
                                    gagaga = 0
                            else: #If school subject is not 0, add school subject in MongoDB
                                remove_things_in_front = school_subject.split(' ', 1)[1]
                                find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                                find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                                array_document_school_time_table = find_document_school_time_table[0]["_id"]
                                array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                                if remove_things_in_front == "MISURE ELETTRICHE":
                                    remove_things_in_front = "MISURE"
                                elif remove_things_in_front == " EDUCAZIONE ATTIVITA' MOTORIE":
                                    remove_things_in_front = "MOTORIA"
                                elif remove_things_in_front == "EDUCAZIONE ATTIVITA' MOTORIE":
                                    remove_things_in_front = "Motoria"
                                elif remove_things_in_front == "DIRITTO ED ECONOMIA":
                                    remove_things_in_front = "DIRITTO"
                                elif remove_things_in_front == "INGLESE PROFESSIONALE":
                                    remove_things_in_front = "INGLESE PRO."
                                elif remove_things_in_front == "LABORATORIO ELETTRICO":
                                    remove_things_in_front = "LAB. ELETTRICO"
                                elif remove_things_in_front == "LINGUA INGLESE":
                                    remove_things_in_front = "INGLESE"
                                elif remove_things_in_front == "SCIENZE INTEGRATE - FISICA":
                                    remove_things_in_front = "FISICA"

                                # Add school subject in MongoDB beacause school subject is not 0
                                collection.update_one(
                                    { "_id": ObjectId(array_document_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga) + ".Subject": remove_things_in_front,
                                    }
                                }
                            )
                                collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga) + ".Subject": remove_things_in_front,
                                    }
                                }
                            )
                                number += 1
                                gagaga += 1
                                dont_repeat += 1
                                if number == 9:
                                    number = 1
                                if gagaga == 8:
                                    gagaga = 0
                        else:
                            datetime_obj = datetime.datetime.strptime(day, "%Y-%m-%d %H:%M:%S").strftime("%d %m %Y")
                            convert_date_to_day = datetime.datetime.strptime(datetime_obj, '%d %m %Y').strftime('%A')
                            array_test = []
                            array_test.append(convert_date_to_day)
                            number_day += 1
                            if school_subject == 0: #If school subject is 0, add "" in MongoDB
                                find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                                find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                                array_document_school_time_table = find_document_school_time_table[0]["_id"]
                                array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                                collection.update_one(
                                    { "_id": ObjectId(array_document_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga) + ".Subject": remove_things_in_front,
                                    }
                                }
                            )
                                collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga) + ".Subject": school_subject,
                                    }
                                }
                            )

                                number += 1
                                gagaga += 1
                                dont_repeat += 1
                                number = 0
                                if number == 9:
                                    number = 1
                                if gagaga == 8:
                                    gagaga = 0
                            else: #If school subject is not 0, add school subject in MongoDB
                                if school_subject == "21PM1 MISURE ELETTRICHE":
                                    school_subject = "MISURE"
                                elif school_subject == " CEAM  EDUCAZIONE ATTIVITA' MOTORIE":
                                    school_subject = "MOTORIA"
                                elif school_subject == "CEAM  EDUCAZIONE ATTIVITA' MOTORIE":
                                    school_subject = "MOTORIA"
                                elif school_subject == "CSGGE1 DIRITTO ED ECONOMIA":
                                    school_subject = "DIRITTO"
                                elif school_subject == "PR1 INGLESE PROFESSIONALE":
                                    school_subject = "INGLESE PRO."
                                elif school_subject == "PR1 LABORATORIO ELETTRICO":
                                    school_subject = "LAB. ELETTRICO"
                                elif school_subject == "CLING LINGUA INGLESE":
                                    school_subject = "INGLESE"
                                elif school_subject == "CMST2 SCIENZE INTEGRATE - FISICA":
                                    school_subject = "FISICA"
                                elif school_subject == "PR1 DISEGNO":
                                    school_subject = "DISEGNO"
                                elif school_subject == "CSGGE2 STORIA":
                                    school_subject = "STORIA"
                                elif school_subject == "CMST1 MATEMATICA":
                                    school_subject = "MATEMATICA"

                                find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                                find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                                array_document_school_time_table = find_document_school_time_table[0]["_id"]
                                array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                                # Add school subject in MongoDB beacause school subject is not 0
                                collection.update_one(
                                    { "_id": ObjectId(array_document_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga)+ ".Subject": school_subject,
                                    }
                                }
                            )
                                collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga)+ ".Subject": school_subject,
                                    }
                                }
                            )
                                number += 1
                                gagaga += 1
                                dont_repeat += 1
                                if number == 9:
                                    number = 1
                                if gagaga == 8:
                                    gagaga = 0

                    #Search teacher
                    if dont_repeat_teacher == 9:
                        check_repeat_teacher += 1 # 13
                        if check_repeat_teacher == 5:
                            check_repeat_teacher = 0
                            dont_repeat_teacher = 0
                            #number = 1
                    else:
                        if day == "None":
                            if teacher == 0:
                                number_teacher += 1
                                gagaga_teacher += 1
                                dont_repeat_teacher += 1
                                if number_teacher == 9:
                                    number_teacher = 1
                                if gagaga_teacher == 9:
                                    gagaga_teacher = 0
                            else:
                                find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                                find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                                array_document_school_time_table = find_document_school_time_table[0]["_id"]
                                array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                                collection.update_one(
                                    { "_id": ObjectId(array_document_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_teacher) + ".Teacher": teacher,
                                    }
                                }
                            )
                                collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_teacher)+ ".Teacher": teacher,
                                    }
                                }
                            )
                                number_teacher += 1
                                gagaga_teacher += 1
                                dont_repeat_teacher += 1
                                if number_teacher == 9:
                                    number_teacher = 1
                                if gagaga_teacher == 8:
                                    gagaga_teacher = 0
                        else:
                            datetime_obj = datetime.datetime.strptime(day, "%Y-%m-%d %H:%M:%S").strftime("%d %m %Y")
                            convert_date_to_day = datetime.datetime.strptime(datetime_obj, '%d %m %Y').strftime('%A')
                            array_test = []
                            array_test.append(convert_date_to_day)
                            find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                            find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                            array_document_school_time_table = find_document_school_time_table[0]["_id"]
                            array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                            collection.update_one(
                                        { "_id": ObjectId(array_document_school_time_table)},
                                        { "$set": {
                                            "School Subject." + array_test[0] + "." + str(gagaga_teacher)+ ".Teacher": teacher,
                                        }
                                    }
                                )
                            collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_teacher)+ ".Teacher": teacher,
                                    }
                                }
                            )
                            number_teacher += 1
                            gagaga_teacher += 1
                            dont_repeat_teacher += 1
                            if number_teacher == 9:
                                number_teacher = 1
                            if gagaga_teacher == 8:
                                gagaga_teacher = 0

                    #Search room school
                    if dont_repeat_room == 9:
                        check_repeat_room += 1
                        if check_repeat_room == 5:
                            check_repeat_room = 0
                            dont_repeat_room = 0
                    else:
                        if day == "None":
                            if room == 0:
                                number_room += 1
                                gagaga_room += 1
                                dont_repeat_room += 1
                                if number_room == 9:
                                    number_room = 1
                                if gagaga_room == 9:
                                    gagaga_room = 0
                            else:
                                find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                                find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                                array_document_school_time_table = find_document_school_time_table[0]["_id"]
                                array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                                collection.update_one(
                                    { "_id": ObjectId(array_document_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_room) + ".Room": room,
                                    }
                                }
                            )
                                collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_room)+ ".Room": room,
                                    }
                                }
                            )
                                number_room += 1
                                gagaga_room += 1
                                dont_repeat_room += 1
                                if number_room == 9:
                                    number_room = 1
                                if gagaga_room == 8:
                                    gagaga_room = 0
                        else:
                            datetime_obj = datetime.datetime.strptime(day, "%Y-%m-%d %H:%M:%S").strftime("%d %m %Y")
                            convert_date_to_day = datetime.datetime.strptime(datetime_obj, '%d %m %Y').strftime('%A')
                            array_test = []
                            array_test.append(convert_date_to_day)
                            find_document_school_time_table = list(collection.find({}, {"Date": long_date}))
                            find_document_archive_school_time_table = list(collection_archive.find({}, {"Date": long_date}))
                            array_document_school_time_table = find_document_school_time_table[0]["_id"]
                            array_document_archive_school_time_table = find_document_archive_school_time_table[0]["_id"]
                            collection.update_one(
                                        { "_id": ObjectId(array_document_school_time_table)},
                                        { "$set": {
                                            "School Subject." + array_test[0] + "." + str(gagaga_room)+ ".Room": room,
                                        }
                                    }
                                )
                            collection_archive.update_one(
                                    { "_id": ObjectId(array_document_archive_school_time_table)},
                                    { "$set": {
                                        "School Subject." + array_test[0] + "." + str(gagaga_room)+ ".Room": room,
                                    }
                                }
                            )
                            number_room += 1
                            gagaga_room += 1
                            dont_repeat_room += 1
                            if number_room == 9:
                                number_room = 1
                            if gagaga_room == 8:
                                gagaga_room = 0

                    # Look for other classes doing PE at the same time as us
                    if school_subject == "CEAM  EDUCAZIONE ATTIVITA' MOTORIE" or school_subject == "CEAM EDUCAZIONE ATTIVITA' MOTORIE":
                        for c in range(1,100):
                            search_motoria = ws.cell(row=i, column=c).value
                            #print(search_motoria)
                            if search_motoria == "CEAM  EDUCAZIONE ATTIVITA' MOTORIE" or search_motoria == "CEAM EDUCAZIONE ATTIVITA' MOTORIE":
                                if c == column:
                                    pass
                                else:
                                    search_class = ws.cell(row=3, column=c).value
                                    collection.update_one(
                                        { "_id": ObjectId(array_document_school_time_table)},
                                        { "$set": {
                                            "School Subject." + array_test[0] + "." + str(gagaga_room-1)+ ".PE with": search_class,
                                        }
                                    }
                                )
                            else:
                                pass

                    # Search if class le3 is busy
                    for c in range(1,100):
                        search_other_subject = ws.cell(row=i, column=c).value
                        search_room = ws.cell(row=i, column=c).value
                        if c == column:
                            pass
                        else:
                            if teacher == "Martin":
                                if search_other_subject == "le3" or search_other_subject == "le":
                                    search_class = ws.cell(row=3, column=c-2).value
                                    if search_class == "2elci":
                                        pass
                                    else:
                                        if c == column:
                                            pass
                                        else:
                                            if search_other_subject == "le3":
                                                collection.update_one(
                                                    { "_id": ObjectId(array_document_school_time_table)},
                                                    { "$set": {
                                                        "School Subject." + array_test[0] + "." + str(gagaga_room-1)+ ".le3 is busy": search_class,
                                                    }
                                                }
                                            )
                                            if search_other_subject == "le":
                                                collection.update_one(
                                                    { "_id": ObjectId(array_document_school_time_table)},
                                                    { "$set": {
                                                        "School Subject." + array_test[0] + "." + str(gagaga_room-1)+ ".le is busy": search_class,
                                                    }
                                                }
                                            )
                                else:
                                    pass

update_time_school()