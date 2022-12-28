from dotenv import load_dotenv
import openpyxl as xl
import urllib
import pymongo
import os
import datetime
from bson.objectid import ObjectId
import os

load_dotenv() #Load .env file
PASSWORD_MONGODB = os.getenv('PASSWORD_MONGODB') #Password for MongoDB
URL_MONGODB = os.getenv('URL_MONGODB') #URL for MongoDB
mongo_url = "mongodb+srv://elci:" + urllib.parse.quote_plus(PASSWORD_MONGODB) + URL_MONGODB #URL for MongoDB (with password)
client = pymongo.MongoClient(mongo_url) #Connect to MongoDB
database = client["website-class"] #Database name
collection = database["school-time-table"]
collection_archive = database["archive-school-time-table"]

x = collection.delete_many({})

#using read_excel() method to read our excel file and storing the same in the variable named "df "
workbook = xl.load_workbook(filename="test.xlsx")

ws = workbook.active

current_time = datetime.datetime.now()
day = str(current_time.day)
month = str(current_time.month)
year = str(current_time.year)
hour = str(current_time.hour)
minute = str(current_time.minute)
long_date = day + "-" + month + "-" + year + " " + hour + ":" + minute
mydict = { 
    "Date": long_date,
    "School Subject": [],
    "Teacher": [],
}
x = collection.insert_one(mydict)
x = collection_archive.insert_one(mydict)

for row in range (1, 100):
    # column B ~ column F
    for column in range (1, 100):
        cell = ws.cell(row, column)
        if cell.value == "2elci":
            ws.cell(row=cell.row, column=column).value
            #Search school time table
            for i in range(4,80):
                school_subject = ws.cell(row=i, column=column).value
                if school_subject == 0:
                    find_document_username = list(collection.find({}, {"Date": long_date}))
                    array_username = find_document_username[0]["_id"]
                    collection.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "School Subject": "null" }
                            }
                        )
                    collection_archive.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "School Subject": "null" }
                            }
                        )
                else:
                    remove_things_in_front = school_subject.split(' ', 1)[1]
                    find_document_username = list(collection.find({}, {"Date": long_date}))
                    array_username = find_document_username[0]["_id"]
                    collection.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "School Subject": str(remove_things_in_front) }
                            }
                        )
                    collection_archive.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "School Subject": str(remove_things_in_front) }
                            }
                        )
                        
            #Search teacher
            for i in range(4, 80):
                teacher = ws.cell(row=i, column=column+1).value
                column = column
                if teacher == 0:
                    find_document_username = list(collection.find({}, {"Date": long_date}))
                    array_username = find_document_username[0]["_id"]
                    collection.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "Teacher": "null" }
                            }
                        )
                    collection_archive.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "Teacher": teacher }
                            }
                        )
                else:
                    find_document_username = list(collection.find({}, {"Date": long_date}))
                    array_username = find_document_username[0]["_id"]
                    collection.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "Teacher": teacher }
                            }
                        )
                    collection_archive.update_one(
                        { "_id": ObjectId(array_username)},
                            {
                                "$push": { "Teacher": teacher }
                            }
                        )